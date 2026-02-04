from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import json
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


KV_HEADERS = ["Key", "Value", "Unit", "Type", "Description", "Required", "Example"]
CASE_HEADERS = ["case_label", "key", "value", "type", "description"]
BEARING_TABLE_HEADERS = ["bearing_id", "k1", "k2", "k3", "k4", "k5", "k6"]


def _style_header(ws, headers: List[str]):
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 70)


def _parse_typed_value(raw: Any, typ: str | None) -> Any:
    if raw is None:
        return None
    if typ is None or str(typ).strip() == "":
        # leave as-is (openpyxl already gives numeric types)
        return raw

    t = str(typ).strip().lower()

    if t in {"str", "string"}:
        return str(raw)

    if t in {"int", "integer"}:
        return int(float(raw))

    if t in {"float", "number"}:
        return float(raw)

    if t in {"bool", "boolean"}:
        if isinstance(raw, bool):
            return raw
        s = str(raw).strip().lower()
        return s in {"1", "true", "t", "yes", "y"}

    # list handling: accept JSON list or comma-separated
    if t.startswith("list"):
        if isinstance(raw, list):
            return raw
        s = str(raw).strip()
        if s == "":
            return []
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                return obj
        except Exception:
            pass
        parts = [p.strip() for p in s.split(",") if p.strip() != ""]
        if t in {"list[int]", "list[integer]"}:
            return [int(float(x)) for x in parts]
        if t in {"list[float]", "list[number]"}:
            return [float(x) for x in parts]
        return parts

    if t in {"json", "dict"}:
        return json.loads(str(raw))

    return raw


def derive_section_from_dimensions(params: Dict[str, Any]) -> None:
    """Derive section convenience variables (b1.., h1.., Ag) from PSCI dimension inputs.

    If the upstream code already provides these values, this will overwrite them to keep
    everything consistent whenever the user edits UF/UT/etc in Excel.
    """
    need = ["UF", "UT", "UFT", "WH", "WT", "LT", "LF", "LFT"]
    if not all(k in params for k in need):
        return

    UF = float(params["UF"])
    UT = float(params["UT"])
    UFT = float(params["UFT"])
    WH = float(params["WH"])
    WT = float(params["WT"])
    LT = float(params["LT"])
    LF = float(params["LF"])
    LFT = float(params["LFT"])

    # widths (mm)
    b1 = UF
    b2 = (UF - WT) / 2.0
    b3 = WT
    b4 = (LF - WT) / 2.0
    b5 = LF

    # heights (mm)
    h1 = UT
    h2 = UFT
    h3 = UFT + WH + LFT
    h4 = LFT
    h5 = LT

    # gross area (mm^2)
    Ag = b1 * h1 + b2 * h2 + b3 * h3 + b4 * h4 + b5 * h5

    params.update(
        dict(
            b1=b1, b2=b2, b3=b3, b4=b4, b5=b5,
            h1=h1, h2=h2, h3=h3, h4=h4, h5=h5,
            Ag=Ag,
        )
    )


def derive_bearing_stiffness(params: Dict[str, Any], bearings_table: Optional[List[List[float]]] = None) -> None:
    """Update Bearing_Stiffness based on 'bearing_mode'.

    - bearing_mode == 'multiplier' (default): multiply the baseline Bearing_Stiffness by bearing_multiplier
    - bearing_mode == 'table'      : replace Bearing_Stiffness with values from BearingsTable sheet
    """
    mode = str(params.get("bearing_mode", "multiplier")).strip().lower()

    if mode == "table" and bearings_table is not None:
        params["Bearing_Stiffness"] = bearings_table
        return

    # multiplier mode
    if "Bearing_Stiffness" in params and params.get("bearing_multiplier") is not None:
        m = float(params["bearing_multiplier"])
        base = np.array(params["Bearing_Stiffness"], dtype=float)
        params["Bearing_Stiffness"] = (base * m).tolist()


def _read_kv_sheet(ws) -> Dict[str, Any]:
    # Expect KV_HEADERS in row 1
    key_col = 1
    val_col = 2
    type_col = 4
    out: Dict[str, Any] = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, key_col).value
        if key is None or str(key).strip() == "":
            continue
        raw_val = ws.cell(r, val_col).value
        typ = ws.cell(r, type_col).value
        val = _parse_typed_value(raw_val, typ)
        if val is None:
            continue
        out[str(key).strip()] = val
    return out


def _read_bearings_table(ws) -> List[List[float]]:
    # Expect BEARING_TABLE_HEADERS in row 1
    rows: List[List[float]] = []
    for r in range(2, ws.max_row + 1):
        bid = ws.cell(r, 1).value
        if bid is None or str(bid).strip() == "":
            continue
        k = []
        for c in range(2, 8):
            v = ws.cell(r, c).value
            k.append(float(v) if v is not None else 0.0)
        rows.append(k)
    return rows


def load_params_from_excel(excel_path: str | Path, case: str = "baseline") -> Dict[str, Any]:
    """Load a parameter dict from the Excel template (v3 multi-sheet or v2 single-sheet).

    Parameters
    ----------
    excel_path:
        Path to workbook
    case:
        Case label from the 'Cases' sheet (default: baseline)

    Returns
    -------
    dict
        Parameter dict that can be passed to builder/modal/moving-load.
    """
    from ..config import make_params

    excel_path = Path(excel_path)
    wb = load_workbook(excel_path, data_only=True)

    # Start from programmatic defaults (ensures missing keys still exist)
    params = make_params(case="baseline")

    # Detect v3 template (presence of Geometry sheet)
    is_v3 = "Geometry" in wb.sheetnames and "Cases" in wb.sheetnames

    # Merge base sheets
    if is_v3:
        kv_sheets = [s for s in wb.sheetnames if s in {
            "Meta","Geometry","Section","Materials","Tendon","Bearings","Modal","Dynamic","Vehicle","Output","Advanced"
        }]
        for s in kv_sheets:
            params.update(_read_kv_sheet(wb[s]))

        bearings_table = None
        if "BearingsTable" in wb.sheetnames:
            bearings_table = _read_bearings_table(wb["BearingsTable"])

        # Cases: apply overrides for the chosen case
        case = str(case).strip().lower()
        if "Cases" in wb.sheetnames:
            overrides_by_case: Dict[str, Dict[str, Any]] = {}
            ws = wb["Cases"]
            # columns: case_label, key, value, type, description
            for r in range(2, ws.max_row + 1):
                case_label = ws.cell(r, 1).value
                key = ws.cell(r, 2).value
                if case_label is None or key is None:
                    continue
                case_label = str(case_label).strip().lower()
                raw_val = ws.cell(r, 3).value
                typ = ws.cell(r, 4).value
                val = _parse_typed_value(raw_val, typ)
                overrides_by_case.setdefault(case_label, {})[str(key).strip()] = val

            if case in overrides_by_case:
                params.update(overrides_by_case[case])
            elif case not in {"baseline", "case1"}:
                raise ValueError(f"Case '{case}' not found in Excel 'Cases' sheet.")

        # Derived values
        derive_section_from_dimensions(params)
        derive_bearing_stiffness(params, bearings_table=bearings_table)
        return params

    # v2 fallback: single-sheet Inputs (+ Cases)
    if "Inputs" in wb.sheetnames:
        params.update(_read_kv_sheet(wb["Inputs"]))

    if "Cases" in wb.sheetnames:
        ws = wb["Cases"]
        case = str(case).strip().lower()
        overrides_by_case: Dict[str, Dict[str, Any]] = {}
        for r in range(2, ws.max_row + 1):
            case_label = ws.cell(r, 1).value
            key = ws.cell(r, 2).value
            if case_label is None or key is None:
                continue
            case_label = str(case_label).strip().lower()
            raw_val = ws.cell(r, 3).value
            typ = ws.cell(r, 4).value
            val = _parse_typed_value(raw_val, typ)
            overrides_by_case.setdefault(case_label, {})[str(key).strip()] = val
        if case in overrides_by_case:
            params.update(overrides_by_case[case])

    derive_section_from_dimensions(params)
    derive_bearing_stiffness(params, bearings_table=None)
    return params


def create_excel_template(excel_path: str | Path) -> Path:
    """Create the **v3 multi-sheet** Excel template.

    The template starts from `bridge_psci.config.make_params('baseline')` defaults, so it will
    remain in-sync with your packaged notebook model.
    """
    from ..config import make_params

    excel_path = Path(excel_path)
    params = make_params(case="baseline")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # README sheet
    ws = wb.create_sheet("README")
    ws["A1"] = "Bridge PSCI Excel Template (v3)"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "1) Fill sheets: Geometry / Section / Materials / Tendon / Bearings / Modal / Dynamic"
    ws["A4"] = "2) Optional: edit BearingsTable (bearing_mode='table') or use bearing_multiplier (default)."
    ws["A5"] = "3) Add scenario overrides in 'Cases' sheet (case_label, key, value, type)."
    ws["A6"] = "4) Run: python scripts/run_excel.py --excel bridge_input_template.xlsx --case baseline"
    ws.column_dimensions["A"].width = 110

    def add_kv_sheet(name: str, rows: List[Tuple[str, Any, str, str, str, str, str]]):
        ws = wb.create_sheet(name)
        _style_header(ws, KV_HEADERS)
        for row in rows:
            ws.append(list(row))
        ws.freeze_panes = "A2"
        _autosize(ws)

    # Helper to build rows
    def row(key, unit="", typ="", desc="", required="Y", example=""):
        val = params.get(key, "")
        # openpyxl can't store lists/dicts directly -> write JSON string
        if isinstance(val, (list, dict)):
            val = json.dumps(val)
        return (key, val, unit, typ, desc, required, example)

    add_kv_sheet("Meta", [
        ("project_name", "PSCI Bridge", "", "str", "Project label used for output folders", "N", "HyojaBridge_UP"),
        ("notes", "", "", "str", "Free-form notes", "N", ""),
    ])

    add_kv_sheet("Geometry", [
        row("Bridge_width", "m", "float", "Overall bridge width", "Y", "12.5"),
        row("Bridge_skew", "deg", "float", "Skew angle", "N", "0"),
        row("girder_number", "ea", "int", "Number of PSCI girders", "Y", "6"),
        row("girder_spacing", "m", "float", "Center-to-center girder spacing", "Y", "1.0"),
        row("Left_Cantilever", "m", "float", "Left cantilever width", "Y", "1.25"),
        row("Right_Cantilever", "m", "float", "Right cantilever width", "Y", "1.25"),
        row("girder_length", "mm", "float", "Girder design length (legacy units)", "Y", "29700"),
        row("girder_H", "mm", "float", "Girder depth", "Y", "2450"),
        row("gravity", "m/s^2", "float", "Gravity (negative downward)", "Y", "-9.81"),
    ])

    add_kv_sheet("Section", [
        row("UF", "mm", "float", "Upper flange width", "Y", "700"),
        row("UT", "mm", "float", "Upper flange thickness", "Y", "150"),
        row("UFT", "mm", "float", "Upper fillet thickness", "Y", "50"),
        row("WH", "mm", "float", "Web height", "Y", "1700"),
        row("WT", "mm", "float", "Web thickness", "Y", "200"),
        row("LFT", "mm", "float", "Lower fillet thickness", "Y", "50"),
        row("LT", "mm", "float", "Lower flange thickness", "Y", "150"),
        row("LF", "mm", "float", "Lower flange width", "Y", "800"),
        # Derived (auto)
        ("b1", "", "", "float", "Derived (auto) — do not edit", "N", ""),
        ("h1", "", "", "float", "Derived (auto) — do not edit", "N", ""),
        ("Ag", "mm^2", "float", "Derived (auto) gross area", "N", ""),
    ])

    add_kv_sheet("Materials", [
        row("Ec", "MPa", "float", "Concrete elastic modulus", "Y", "24000"),
        row("E_girder1", "MPa", "list[float]", "Girder E list (length = 2*girder_number)", "Y", "[28000,28000,...]"),
        row("E_girder2", "MPa", "list[float]", "Girder E list (length = 2*girder_number)", "Y", "[28000,28000,...]"),
        row("E_deck1", "MPa", "list[float]", "Deck E list (length = girder_number+1)", "Y", "[25000,...]"),
        row("E_deck2", "MPa", "list[float]", "Deck E list (length = girder_number+1)", "Y", "[25000,...]"),
        row("diaphragm1_Ec", "MPa", "list[float]", "Diaphragm E list (1)", "N", ""),
        row("diaphragm2_Ec", "MPa", "list[float]", "Diaphragm E list (2)", "N", ""),
    ])

    add_kv_sheet("Tendon", [
        row("number_tendon", "ea", "int", "Number of tendon groups", "Y", "4"),
        row("tendon_horizontal_length", "mm", "float", "Tendon horizontal length", "Y", "28000"),
        row("y_intercept_list", "mm", "list[float]", "Tendon y-intercepts", "Y", "[... ]"),
        row("z_intercept_list", "mm", "list[float]", "Tendon z-intercepts", "Y", "[... ]"),
        row("Ap_N", "mm^2", "list[float]", "Tendon areas per group", "Y", "[... ]"),
        row("PE", "MPa", "float", "Prestress parameter used in notebook", "N", "600"),
    ])

    add_kv_sheet("Bearings", [
        ("bearing_mode", "multiplier/table", "str", "How to set Bearing_Stiffness", "Y", "multiplier"),
        row("bearing_multiplier", "-", "float", "Multiplier applied in 'multiplier' mode", "Y", "0.3"),
        ("Bearing_Stiffness", "N/mm", "json", "Baseline stiffness table (auto); use BearingsTable to override", "N", ""),
    ])

    add_kv_sheet("Modal", [
        row("numEigen", "ea", "int", "Number of eigenvalues/modes", "Y", "3"),
        row("zeta", "-", "float", "Modal damping ratio", "Y", "0.015"),
    ])

    add_kv_sheet("Dynamic", [
        row("dt", "s", "float", "Time step", "Y", "0.1"),
        row("velocity_kmh", "km/h", "float", "Vehicle speed", "Y", "10"),
        row("load", "N", "float", "Vehicle axle/load parameter used in notebook", "N", ""),
        row("pave_thick", "mm", "list[float]", "Pavement thickness", "N", "[80]"),
    ])

    add_kv_sheet("Output", [
        row("girder3_start_tag", "", "int", "Node tag start for girder3 (legacy)", "N", "3001"),
        row("girder4_start_tag", "", "int", "Node tag start for girder4 (legacy)", "N", "4001"),
        row("girder_n_nodes", "", "int", "Nodes per girder centroid line (legacy)", "N", "149"),
    ])

    # Advanced sheet for any extra keys (keeps full coverage)
    known = set()
    for s in ["Geometry","Section","Materials","Tendon","Bearings","Modal","Dynamic","Output"]:
        for r in wb[s].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                known.add(str(r[0]).strip())

    advanced_rows = []
    for k in sorted(params.keys()):
        if k in known:
            continue
        # Exclude internal notes keys
        if k in {"project_name","notes"}:
            continue
        val = params.get(k, "")
        if isinstance(val, (list, dict)):
            val = json.dumps(val)
        advanced_rows.append((k, val, "", "", "", "N", ""))

    add_kv_sheet("Advanced", advanced_rows[:200])  # keep workbook light; you can add more later

    # Cases sheet
    ws = wb.create_sheet("Cases")
    _style_header(ws, CASE_HEADERS)
    ws.append(["baseline", "bearing_multiplier", 0.3, "float", "Example override"])
    ws.freeze_panes = "A2"
    _autosize(ws)

    # BearingsTable sheet (optional)
    ws = wb.create_sheet("BearingsTable")
    _style_header(ws, BEARING_TABLE_HEADERS)
    # Provide default ordering (12 bearings: A1_B1..B6, A2_B1..B6)
    default_b = np.array(params.get("Bearing_Stiffness"), dtype=float)
    bearing_ids = [f"A1_B{i}" for i in range(1,7)] + [f"A2_B{i}" for i in range(1,7)]
    for i,bid in enumerate(bearing_ids):
        rowvals = [bid] + [float(x) for x in default_b[i].tolist()]
        ws.append(rowvals)
    ws.freeze_panes = "A2"
    _autosize(ws)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    return excel_path
